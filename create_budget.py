"""
Budget Master Spreadsheet Creator
Creates a comprehensive budget management Excel workbook
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime, timedelta
from openpyxl.formatting.rule import FormulaRule

# Create workbook
wb = Workbook()

# Styles
header_font = Font(bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
subheader_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
money_format = '"$"#,##0.00'
percent_format = '0.0%'
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def style_header(cell):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

def style_cell(cell, is_money=False, is_percent=False):
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center' if not is_money else 'right', vertical='center')
    if is_money:
        cell.number_format = money_format
    if is_percent:
        cell.number_format = percent_format

# ============================================
# SHEET 1: DASHBOARD
# ============================================
ws = wb.active
ws.title = "Dashboard"

# Title
ws.merge_cells('A1:H1')
ws['A1'] = "💰 BUDGET DASHBOARD - Joshua's Financial Command Center"
ws['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws['A1'].alignment = Alignment(horizontal='center')

# Income Summary Section
ws['A3'] = "INCOME SUMMARY"
ws['A3'].font = Font(bold=True, size=14)
ws.merge_cells('A3:D3')

headers = ['Category', 'Annual', 'Monthly', 'Per Paycheck']
for i, h in enumerate(headers, 1):
    cell = ws.cell(row=4, column=i, value=h)
    style_header(cell)

income_data = [
    ['Gross Salary', 76000, '=B5/12', '=B5/26'],
    ['Net Pay (After Deductions)', '=C6*12', 4160, 1920],
]
for r, row_data in enumerate(income_data, 5):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        style_cell(cell, is_money=(c > 1))

# Paycheck Deductions (already taken out)
ws['A8'] = "AUTOMATIC PAYCHECK DEDUCTIONS (Already Deducted)"
ws['A8'].font = Font(bold=True, size=14)
ws.merge_cells('A8:D8')

for i, h in enumerate(['Deduction', 'Per Paycheck', 'Monthly', 'Annual'], 1):
    cell = ws.cell(row=9, column=i, value=h)
    style_header(cell)

deductions = [
    ['Roth 401(k) - Your 8%', 253.33, '=B10*26/12', '=B10*26'],
    ['HSA', 126.67, '=B11*26/12', '=B11*26'],
    ['Fed Withholding', 319.22, '=B12*26/12', '=B12*26'],
    ['Fed Med/OASDI', 238.94, '=B13*26/12', '=B13*26'],
    ['CA Withholding', 136.15, '=B14*26/12', '=B14*26'],
    ['Disability Ins', 12.94, '=B15*26/12', '=B15*26'],
    ['TOTAL DEDUCTIONS', '=SUM(B10:B15)', '=SUM(C10:C15)', '=SUM(D10:D15)'],
]
for r, row_data in enumerate(deductions, 10):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        style_cell(cell, is_money=(c > 1))
        if r == 16:
            cell.font = Font(bold=True)

# FREE MONEY Section
ws['A18'] = "🎁 FREE MONEY (Employer Contributions)"
ws['A18'].font = Font(bold=True, size=14, color="228B22")
ws.merge_cells('A18:D18')

for i, h in enumerate(['Benefit', 'Per Paycheck', 'Monthly', 'Annual'], 1):
    cell = ws.cell(row=19, column=i, value=h)
    style_header(cell)

ws.cell(row=20, column=1, value="Employer 401(k) Match (8%)")
ws.cell(row=20, column=2, value='=76000*0.08/26')
ws.cell(row=20, column=3, value='=B20*26/12')
ws.cell(row=20, column=4, value='=B20*26')
for c in range(1, 5):
    cell = ws.cell(row=20, column=c)
    style_cell(cell, is_money=(c > 1))
    cell.fill = green_fill

# Key Financial Stats
ws['F3'] = "KEY STATS"
ws['F3'].font = Font(bold=True, size=14)
ws.merge_cells('F3:H3')

stats = [
    ['Total Retirement Savings/Year', '=D10+D11+D20', '(401k + HSA + Employer)'],
    ['Retirement % of Gross', '=F4/76000', ''],
    ['Effective Tax Rate', '=(D12+D14)/76000', ''],
    ['Take-Home Rate', '=C6*12/76000', ''],
]
for r, (label, val, note) in enumerate(stats, 4):
    ws.cell(row=r, column=6, value=label).font = Font(bold=True)
    cell = ws.cell(row=r, column=7, value=val)
    if r == 5 or r == 6 or r == 7:
        style_cell(cell, is_percent=True)
    else:
        style_cell(cell, is_money=True)
    ws.cell(row=r, column=8, value=note).font = Font(italic=True, color="666666")

# Column widths
ws.column_dimensions['A'].width = 35
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 15
ws.column_dimensions['F'].width = 30
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 25

# ============================================
# SHEET 2: MONTHLY BUDGET
# ============================================
ws2 = wb.create_sheet("Monthly Budget")

ws2['A1'] = "MONTHLY BUDGET PLANNER"
ws2['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws2.merge_cells('A1:E1')

# Income Section
ws2['A3'] = "MONTHLY INCOME"
ws2['A3'].font = Font(bold=True, size=12)
style_header(ws2['A3'])
ws2.merge_cells('A3:C3')

ws2['A4'] = "Net Monthly Income (avg)"
ws2['B4'] = 4160
ws2['C4'] = "Based on $1,920 x 26 / 12"
style_cell(ws2['B4'], is_money=True)

# Fixed Expenses
ws2['A6'] = "FIXED EXPENSES (Non-Negotiable)"
ws2['A6'].font = Font(bold=True, size=12)
style_header(ws2['A6'])
ws2.merge_cells('A6:C6')

fixed_expenses = [
    ['Rent', 1815, 'Fixed'],
    ['Power', 120, 'Estimate'],
    ['Internet', 51.16, 'Fixed'],
    ['Gas (Utilities)', 50, 'Estimate'],
    ['Groceries', 145, 'Budget'],
    ['Credit Card Payment', 230, '4 months remaining'],
]

for i, h in enumerate(['Expense', 'Amount', 'Notes'], 1):
    cell = ws2.cell(row=7, column=i, value=h)
    cell.fill = subheader_fill
    cell.font = Font(bold=True)
    cell.border = thin_border

for r, (exp, amt, note) in enumerate(fixed_expenses, 8):
    ws2.cell(row=r, column=1, value=exp).border = thin_border
    cell = ws2.cell(row=r, column=2, value=amt)
    style_cell(cell, is_money=True)
    ws2.cell(row=r, column=3, value=note).border = thin_border

ws2['A14'] = "TOTAL FIXED"
ws2['A14'].font = Font(bold=True)
ws2['B14'] = '=SUM(B8:B13)'
style_cell(ws2['B14'], is_money=True)
ws2['B14'].font = Font(bold=True)

# After Fixed
ws2['A16'] = "REMAINING AFTER FIXED"
ws2['A16'].font = Font(bold=True, size=12, color="228B22")
ws2['B16'] = '=B4-B14'
style_cell(ws2['B16'], is_money=True)
ws2['B16'].font = Font(bold=True, size=12)
ws2['B16'].fill = green_fill

# Savings Allocation (from remaining)
ws2['A18'] = "SAVINGS ALLOCATION (From Remaining)"
ws2['A18'].font = Font(bold=True, size=12)
style_header(ws2['A18'])
ws2.merge_cells('A18:D18')

for i, h in enumerate(['Category', 'Monthly', '% of Remaining', 'Purpose'], 1):
    cell = ws2.cell(row=19, column=i, value=h)
    cell.fill = subheader_fill
    cell.font = Font(bold=True)
    cell.border = thin_border

savings = [
    ['Roth IRA', 583.33, '=B20/$B$16', 'MAX IT - House + Retirement'],
    ['Emergency/House Fund', 850, '=B21/$B$16', 'Target: $15k (faster!)'],
    ['Brokerage ($50 SPY/$50 QQQ)', 100, '=B22/$B$16', 'Long-term wealth'],
    ['Fun/Variable Spending', 150, '=B23/$B$16', 'HARD LIMIT'],
    ['Buffer (Unexpected)', 65.51, '=B24/$B$16', 'Peace of mind'],
]

for r, (cat, amt, pct, purpose) in enumerate(savings, 20):
    ws2.cell(row=r, column=1, value=cat).border = thin_border
    cell = ws2.cell(row=r, column=2, value=amt)
    style_cell(cell, is_money=True)
    cell = ws2.cell(row=r, column=3, value=pct)
    style_cell(cell, is_percent=True)
    ws2.cell(row=r, column=4, value=purpose).border = thin_border

ws2['A25'] = "TOTAL ALLOCATED"
ws2['A25'].font = Font(bold=True)
ws2['B25'] = '=SUM(B20:B24)'
style_cell(ws2['B25'], is_money=True)
ws2['B25'].font = Font(bold=True)

# Balance Check
ws2['A27'] = "BALANCE CHECK"
ws2['A27'].font = Font(bold=True, size=12)
ws2['B27'] = '=B16-B25'
style_cell(ws2['B27'], is_money=True)
ws2['C27'] = '← Should be $0 or close to it'

ws2.column_dimensions['A'].width = 35
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 25

# ============================================
# SHEET 3: ROTH IRA TRACKER
# ============================================
ws3 = wb.create_sheet("Roth IRA Tracker")

ws3['A1'] = "🎯 ROTH IRA TRACKER - MAX THAT ROTH!"
ws3['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws3.merge_cells('A1:E1')

# Why Max Roth IRA
ws3['A3'] = "WHY MAXING ROTH IRA IS IMPORTANT:"
ws3['A3'].font = Font(bold=True, size=12)
ws3.merge_cells('A3:E3')

reasons = [
    "✓ Tax-FREE growth forever - never pay taxes on gains",
    "✓ First $10k can go toward house (first-time homebuyer exception)",
    "✓ Can withdraw CONTRIBUTIONS anytime tax/penalty free",
    "✓ $7,000/year limit - USE IT OR LOSE IT (can't make up later)",
    "✓ At your age, compound growth is your superpower",
]
for r, reason in enumerate(reasons, 4):
    ws3.cell(row=r, column=1, value=reason)
    ws3.merge_cells(f'A{r}:E{r}')

# 2025 Progress
ws3['A10'] = "2025 CONTRIBUTION PROGRESS"
ws3['A10'].font = Font(bold=True, size=14)
ws3.merge_cells('A10:E10')

for i, h in enumerate(['Month', 'Contribution', 'YTD Total', 'Remaining', '% Complete'], 1):
    cell = ws3.cell(row=11, column=i, value=h)
    style_header(cell)

months = ['January', 'February', 'March', 'April', 'May', 'June',
          'July', 'August', 'September', 'October', 'November', 'December']

for r, month in enumerate(months, 12):
    ws3.cell(row=r, column=1, value=month).border = thin_border

    # Contribution (enter manually)
    cell = ws3.cell(row=r, column=2, value=583.33 if month == 'January' else 0)
    style_cell(cell, is_money=True)

    # YTD Total
    if r == 12:
        formula = '=B12'
    else:
        formula = f'=C{r-1}+B{r}'
    cell = ws3.cell(row=r, column=3, value=formula)
    style_cell(cell, is_money=True)

    # Remaining
    cell = ws3.cell(row=r, column=4, value=f'=7000-C{r}')
    style_cell(cell, is_money=True)

    # % Complete
    cell = ws3.cell(row=r, column=5, value=f'=C{r}/7000')
    style_cell(cell, is_percent=True)

# Summary
ws3['A25'] = "Target Monthly Contribution:"
ws3['B25'] = 583.33
style_cell(ws3['B25'], is_money=True)
ws3['C25'] = "= $7,000 / 12 months"

ws3['A26'] = "Annual Limit (2025):"
ws3['B26'] = 7000
style_cell(ws3['B26'], is_money=True)

ws3['A27'] = "Your Total Contributed:"
ws3['B27'] = '=C23'
style_cell(ws3['B27'], is_money=True)
ws3['B27'].font = Font(bold=True)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 15
ws3.column_dimensions['D'].width = 15
ws3.column_dimensions['E'].width = 15

# ============================================
# SHEET 4: CREDIT CARD PAYOFF
# ============================================
ws4 = wb.create_sheet("Credit Card Payoff")

ws4['A1'] = "💳 CREDIT CARD DEBT PAYOFF TRACKER"
ws4['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws4.merge_cells('A1:E1')

# Debt Summary
ws4['A3'] = "DEBT SUMMARY"
ws4['A3'].font = Font(bold=True, size=12)
style_header(ws4['A3'])
ws4.merge_cells('A3:C3')

ws4['A4'] = "Total Debt (Enter Here):"
ws4['B4'] = 920  # $230 x 4 months
style_cell(ws4['B4'], is_money=True)
ws4['B4'].fill = yellow_fill

ws4['A5'] = "Monthly Payment:"
ws4['B5'] = 230
style_cell(ws4['B5'], is_money=True)

ws4['A6'] = "Months to Payoff:"
ws4['B6'] = '=CEILING(B4/B5,1)'
style_cell(ws4['B6'])

ws4['A7'] = "Target Payoff Date:"
ws4['B7'] = '=TODAY()+B6*30'
ws4['B7'].number_format = 'MMM YYYY'

# Payment Schedule
ws4['A9'] = "PAYMENT SCHEDULE"
ws4['A9'].font = Font(bold=True, size=12)
style_header(ws4['A9'])
ws4.merge_cells('A9:E9')

for i, h in enumerate(['Month', 'Payment', 'Remaining Balance', 'Paid?', 'Date Paid'], 1):
    cell = ws4.cell(row=10, column=i, value=h)
    cell.fill = subheader_fill
    cell.font = Font(bold=True)
    cell.border = thin_border

for r in range(11, 15):  # 4 months
    ws4.cell(row=r, column=1, value=f'Month {r-10}').border = thin_border
    cell = ws4.cell(row=r, column=2, value=230)
    style_cell(cell, is_money=True)

    if r == 11:
        formula = f'=$B$4-B{r}'
    else:
        formula = f'=C{r-1}-B{r}'
    cell = ws4.cell(row=r, column=3, value=formula)
    style_cell(cell, is_money=True)

    ws4.cell(row=r, column=4, value='☐').border = thin_border  # Checkbox placeholder
    ws4.cell(row=r, column=5, value='').border = thin_border

# After Payoff
ws4['A16'] = "🎉 AFTER PAYOFF - REDIRECT $230/MONTH TO:"
ws4['A16'].font = Font(bold=True, size=12, color="228B22")
ws4.merge_cells('A16:E16')

redirect_options = [
    ['Emergency Fund', 'Reach $15k faster'],
    ['Brokerage', 'More investing power'],
    ['Fun Money', 'Reward yourself!'],
]
for r, (option, note) in enumerate(redirect_options, 17):
    ws4.cell(row=r, column=1, value=f'• {option}')
    ws4.cell(row=r, column=2, value=note).font = Font(italic=True, color="666666")

ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 15
ws4.column_dimensions['C'].width = 20
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 15

# ============================================
# SHEET 5: WORK EXPENSE FLOAT TRACKER
# ============================================
ws5 = wb.create_sheet("Work Expenses")

ws5['A1'] = "🏢 WORK EXPENSE FLOAT TRACKER"
ws5['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws5.merge_cells('A1:G1')

ws5['A3'] = "Track expenses you pay out-of-pocket for work reimbursement"
ws5['A3'].font = Font(italic=True, color="666666")
ws5.merge_cells('A3:G3')

# Summary
ws5['A5'] = "CURRENT FLOAT SUMMARY"
ws5['A5'].font = Font(bold=True, size=12)
style_header(ws5['A5'])
ws5.merge_cells('A5:C5')

ws5['A6'] = "Total Outstanding:"
ws5['B6'] = '=SUMIF(F10:F100,"Pending",D10:D100)'
style_cell(ws5['B6'], is_money=True)
ws5['B6'].fill = yellow_fill

ws5['A7'] = "Expected Reimbursement Date:"
ws5['B7'] = '=MIN(G10:G100)'
ws5['B7'].number_format = 'MM/DD/YYYY'

# Expense Log
for i, h in enumerate(['Date', 'Description', 'Category', 'Amount', 'Receipt?', 'Status', 'Expected Reimb.'], 1):
    cell = ws5.cell(row=9, column=i, value=h)
    style_header(cell)

# Sample entries
sample_expenses = [
    ['01/15/2025', 'Client lunch - Project X', 'Meals', 45.00, 'Yes', 'Pending', '02/01/2025'],
    ['01/18/2025', 'Uber to client site', 'Travel', 28.50, 'Yes', 'Pending', '02/01/2025'],
    ['', '', '', '', '', '', ''],  # Empty rows for future entries
    ['', '', '', '', '', '', ''],
    ['', '', '', '', '', '', ''],
]

for r, expense in enumerate(sample_expenses, 10):
    for c, val in enumerate(expense, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        cell.border = thin_border
        if c == 4 and val:  # Amount column
            cell.number_format = money_format

# Float Impact Analysis
ws5['A17'] = "FLOAT IMPACT ON BUDGET"
ws5['A17'].font = Font(bold=True, size=12)
style_header(ws5['A17'])
ws5.merge_cells('A17:C17')

ws5['A18'] = "Max Safe Float (1 paycheck):"
ws5['B18'] = 1920
style_cell(ws5['B18'], is_money=True)

ws5['A19'] = "Current Float:"
ws5['B19'] = '=B6'
style_cell(ws5['B19'], is_money=True)

ws5['A20'] = "Remaining Float Capacity:"
ws5['B20'] = '=B18-B19'
style_cell(ws5['B20'], is_money=True)

ws5['A22'] = "⚠️ If float > 1 paycheck, delay non-essential spending until reimbursed"
ws5['A22'].font = Font(italic=True, color="CC0000")
ws5.merge_cells('A22:G22')

ws5.column_dimensions['A'].width = 15
ws5.column_dimensions['B'].width = 25
ws5.column_dimensions['C'].width = 12
ws5.column_dimensions['D'].width = 12
ws5.column_dimensions['E'].width = 10
ws5.column_dimensions['F'].width = 12
ws5.column_dimensions['G'].width = 18

# ============================================
# SHEET 6: EMERGENCY FUND TRACKER
# ============================================
ws6 = wb.create_sheet("Emergency Fund")

ws6['A1'] = "🏦 EMERGENCY FUND TRACKER"
ws6['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws6.merge_cells('A1:E1')

ws6['A3'] = "Goal: $15,000 (About 6 months of essential expenses)"
ws6['A3'].font = Font(bold=True, size=12)

# Current Status
ws6['A5'] = "CURRENT STATUS"
style_header(ws6['A5'])
ws6.merge_cells('A5:C5')

ws6['A6'] = "Current Balance:"
ws6['B6'] = 0  # Enter current balance
style_cell(ws6['B6'], is_money=True)
ws6['B6'].fill = yellow_fill

ws6['A7'] = "Target:"
ws6['B7'] = 15000
style_cell(ws6['B7'], is_money=True)

ws6['A8'] = "Remaining to Goal:"
ws6['B8'] = '=B7-B6'
style_cell(ws6['B8'], is_money=True)

ws6['A9'] = "Progress:"
ws6['B9'] = '=B6/B7'
style_cell(ws6['B9'], is_percent=True)

ws6['A11'] = "Monthly Contribution:"
ws6['B11'] = 750
style_cell(ws6['B11'], is_money=True)

ws6['A12'] = "Months to Goal:"
ws6['B12'] = '=CEILING(B8/B11,1)'

ws6['A13'] = "Target Date:"
ws6['B13'] = '=TODAY()+B12*30'
ws6['B13'].number_format = 'MMM YYYY'

# Monthly Progress
ws6['A15'] = "MONTHLY CONTRIBUTIONS"
style_header(ws6['A15'])
ws6.merge_cells('A15:D15')

for i, h in enumerate(['Month', 'Contribution', 'Running Total', '% to Goal'], 1):
    cell = ws6.cell(row=16, column=i, value=h)
    cell.fill = subheader_fill
    cell.font = Font(bold=True)
    cell.border = thin_border

for r in range(17, 37):  # 20 months of tracking
    ws6.cell(row=r, column=1, value='').border = thin_border
    cell = ws6.cell(row=r, column=2, value=0)
    style_cell(cell, is_money=True)

    if r == 17:
        formula = f'=$B$6+B{r}'
    else:
        formula = f'=C{r-1}+B{r}'
    cell = ws6.cell(row=r, column=3, value=formula)
    style_cell(cell, is_money=True)

    cell = ws6.cell(row=r, column=4, value=f'=C{r}/$B$7')
    style_cell(cell, is_percent=True)

ws6.column_dimensions['A'].width = 20
ws6.column_dimensions['B'].width = 15
ws6.column_dimensions['C'].width = 15
ws6.column_dimensions['D'].width = 12

# ============================================
# SHEET 7: PAYCHECK TRACKER
# ============================================
ws7 = wb.create_sheet("Paycheck Tracker")

ws7['A1'] = "📅 PAYCHECK-BY-PAYCHECK TRACKER"
ws7['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws7.merge_cells('A1:H1')

ws7['A3'] = "Track each paycheck and how you allocate it"
ws7['A3'].font = Font(italic=True, color="666666")

# Standard allocation per paycheck
ws7['A5'] = "STANDARD PAYCHECK ALLOCATION (~$1,920 net)"
style_header(ws7['A5'])
ws7.merge_cells('A5:C5')

allocations = [
    ['Fixed Expenses (half monthly)', 1110, '=($1815+$120+$51+$50)/2 + $145/2'],
    ['Roth IRA', 291.67, '$583.33/2 per paycheck'],
    ['Emergency Fund', 425, '$850/2 per paycheck'],
    ['Brokerage', 50, '$100/2 per paycheck'],
    ['Fun Money', 75, '$150/2 per paycheck - HARD LIMIT'],
    ['Buffer/CC if applicable', 0, 'Adjust as needed'],
]

for i, h in enumerate(['Category', 'Amount', 'Notes'], 1):
    cell = ws7.cell(row=6, column=i, value=h)
    cell.fill = subheader_fill
    cell.font = Font(bold=True)
    cell.border = thin_border

for r, (cat, amt, note) in enumerate(allocations, 7):
    ws7.cell(row=r, column=1, value=cat).border = thin_border
    cell = ws7.cell(row=r, column=2, value=amt)
    style_cell(cell, is_money=True)
    ws7.cell(row=r, column=3, value=note).border = thin_border

ws7['A13'] = "TOTAL:"
ws7['B13'] = '=SUM(B7:B12)'
style_cell(ws7['B13'], is_money=True)
ws7['B13'].font = Font(bold=True)

# Actual paycheck log
ws7['A15'] = "ACTUAL PAYCHECK LOG"
style_header(ws7['A15'])
ws7.merge_cells('A15:H15')

headers = ['Pay Date', 'Gross', 'Net', 'Hours', 'Roth IRA', 'E-Fund', 'Brokerage', 'Notes']
for i, h in enumerate(headers, 1):
    cell = ws7.cell(row=16, column=i, value=h)
    cell.fill = subheader_fill
    cell.font = Font(bold=True)
    cell.border = thin_border

# Sample entry
sample = ['01/24/2025', 3250.01, 2162.76, 96, 291.67, 375, 50, 'Extra hours']
for c, val in enumerate(sample, 1):
    cell = ws7.cell(row=17, column=c, value=val)
    cell.border = thin_border
    if c in [2, 3, 5, 6, 7]:
        cell.number_format = money_format

# Empty rows for future entries
for r in range(18, 30):
    for c in range(1, 9):
        ws7.cell(row=r, column=c, value='').border = thin_border

ws7.column_dimensions['A'].width = 12
ws7.column_dimensions['B'].width = 12
ws7.column_dimensions['C'].width = 12
ws7.column_dimensions['D'].width = 8
ws7.column_dimensions['E'].width = 12
ws7.column_dimensions['F'].width = 12
ws7.column_dimensions['G'].width = 12
ws7.column_dimensions['H'].width = 20

# ============================================
# SHEET 8: THE MONEY RULES
# ============================================
ws8 = wb.create_sheet("Money Rules")

ws8['A1'] = "📚 JOSHUA'S MONEY MANAGEMENT RULES"
ws8['A1'].font = Font(bold=True, size=16, color="2E75B6")
ws8.merge_cells('A1:E1')

rules = [
    "",
    "🎯 THE PRIORITY ORDER (Pay Yourself First)",
    "────────────────────────────────────",
    "1. Fixed expenses MUST be covered (rent, utilities, food)",
    "2. Credit card debt gets eliminated (4 months then done!)",
    "3. Roth IRA gets maxed ($583.33/month - tax-free forever)",
    "4. Emergency fund grows ($750/month until $15k)",
    "5. Brokerage for extra wealth building",
    "6. Fun money LAST (but don't skip it - burnout is real)",
    "",
    "💡 KEY INSIGHTS FROM YOUR NUMBERS",
    "────────────────────────────────────",
    "• Your employer gives you FREE 8% 401k match = ~$6,080/year FREE MONEY",
    "• You're already saving 8% in Roth 401k + HSA from paycheck",
    "• Total retirement savings: ~24% of gross (excellent!)",
    "• After CC payoff, you'll have $230 extra/month",
    "",
    "⚠️ WARNING SIGNS TO WATCH",
    "────────────────────────────────────",
    "• Fun spending hitting $308? That's eating into savings",
    "• Work expense float > 1 paycheck? Delay discretionary spending",
    "• Skipping Roth IRA contribution? You lose that year's limit forever",
    "",
    "🏆 YOUR FINANCIAL SUPERPOWERS",
    "────────────────────────────────────",
    "• Young + high savings rate = compound interest machine",
    "• HSA = triple tax advantage (pre-tax in, grows tax-free, tax-free out for medical)",
    "• Roth IRA = flexibility (contributions out anytime, $10k for house)",
    "• Employer match = instant 100% return on investment",
    "",
    "📊 THE MATH THAT MATTERS",
    "────────────────────────────────────",
    "• $583/month in Roth IRA for 30 years @ 7% = ~$700,000",
    "• That $308 fun spending? Over 30 years @ 7% = ~$370,000 opportunity cost",
    "• Every $1 saved in your 20s = ~$7.60 at retirement (7% for 30 years)",
    "",
    "🎮 GAMIFY YOUR FINANCES",
    "────────────────────────────────────",
    "• Set monthly 'high scores' for savings",
    "• Celebrate milestones (first $1k, $5k, $10k in emergency fund)",
    "• Track your net worth monthly - watch it grow!",
]

for r, rule in enumerate(rules, 2):
    ws8.cell(row=r, column=1, value=rule)
    ws8.merge_cells(f'A{r}:E{r}')
    if '🎯' in rule or '💡' in rule or '⚠️' in rule or '🏆' in rule or '📊' in rule or '🎮' in rule:
        ws8.cell(row=r, column=1).font = Font(bold=True, size=12)
    if '────' in rule:
        ws8.cell(row=r, column=1).font = Font(color="AAAAAA")

ws8.column_dimensions['A'].width = 70

# Save workbook
output_path = r"c:\Users\joshua.torres\OneDrive - Sierra-Cedar\Documents\Claude Projects\Budget\Budget_Master.xlsx"
wb.save(output_path)
print("Budget spreadsheet created successfully!")
print(f"Saved to: {output_path}")
print("\nSheets created:")
for sheet in wb.sheetnames:
    print(f"  - {sheet}")
